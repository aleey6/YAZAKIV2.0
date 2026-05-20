"""Génération de fichiers Excel pour l'export des contrats."""

from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import COLORS, fmt_money
from invoice_parser import InvoiceData

def contracts_to_excel(invoice: InvoiceData, plant: str = "", dept: str = "", project: str = "") -> bytes:
    """Génère un fichier Excel avec le détail des contrats."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Détail des contrats"

    # ── En-tête YAZAKI ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    cell_title = ws["A1"]
    cell_title.value = "YAZAKI · Détail des contrats IAM"
    cell_title.font = Font(name="Arial", bold=True, size=14, color=COLORS['blanc'])
    cell_title.fill = PatternFill("solid", fgColor=COLORS['noir'])
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Meta-info
    meta = [
        ("Usine", plant or "—"),
        ("Département", dept or "—"),
        ("Projet", project or "—"),
        ("N° Facture", invoice.invoice_number or "—"),
        ("Date Facture", invoice.invoice_date or "—"),
        ("Période", f"{invoice.period_start or '—'} → {invoice.period_end or '—'}"),
        ("Total (MAD)", f"{invoice.total:,.2f}"),
        ("Nb contrats", str(len(invoice.contracts))),
    ]
    row = 2
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10)
        row += 1

    # ── En-têtes colonnes ───────────────────────────────────────────────────
    headers = ["Page PDF", "Page Doc.", "Type de contrat", "N° d'Appel",
               "Articles Mensuels", "Articles Ponctuels", "Total Contrat (MAD)"]
    header_row = row + 1
    thin = Side(style="thin", color=COLORS['gris_bord'])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=10)
        c.fill = PatternFill("solid", fgColor=COLORS['rouge'])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Données ─────────────────────────────────────────────────────────────
    for i, c_obj in enumerate(invoice.contracts):
        data_row = header_row + 1 + i
        values = [
            c_obj.page_number,
            c_obj.document_page or "—",
            c_obj.contract_type,
            c_obj.phone_number or "—",
            len(c_obj.frais_mensuels),
            len(c_obj.frais_ponctuels),
            c_obj.total_contrat,
        ]
        fill_color = COLORS['gris'] if i % 2 == 0 else COLORS['blanc']
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            if col_idx == 7:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    # ── Ligne Total ─────────────────────────────────────────────────────────
    total_row = header_row + 1 + len(invoice.contracts)
    ws.merge_cells(f"A{total_row}:F{total_row}")
    t = ws.cell(row=total_row, column=1, value="TOTAL")
    t.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=10)
    t.fill = PatternFill("solid", fgColor=COLORS['noir'])
    t.alignment = Alignment(horizontal="right")
    total_val = ws.cell(row=total_row, column=7,
                        value=f"=SUM(G{header_row+1}:G{total_row-1})")
    total_val.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=10)
    total_val.fill = PatternFill("solid", fgColor=COLORS['noir'])
    total_val.number_format = "#,##0.00"
    total_val.alignment = Alignment(horizontal="right")

    # ── Largeurs colonnes ───────────────────────────────────────────────────
    col_widths = [10, 10, 28, 18, 18, 18, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def contracts_to_excel_batch(invoices: list[InvoiceData], plant: str = "", dept: str = "", project: str = "") -> bytes:
    """Génère un fichier Excel avec le détail des contrats pour plusieurs factures."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Détail des contrats (Batch)"

    # ── En-tête YAZAKI ──────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    cell_title = ws["A1"]
    batch_count = len(invoices)
    cell_title.value = f"YAZAKI · Détail des contrats IAM ({batch_count} facture{'s' if batch_count > 1 else ''})"
    cell_title.font = Font(name="Arial", bold=True, size=14, color=COLORS['blanc'])
    cell_title.fill = PatternFill("solid", fgColor=COLORS['noir'])
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Meta-info batch
    meta = [
        ("Usine", plant or "—"),
        ("Département", dept or "—"),
        ("Projet", project or "—"),
        ("Nombre de factures", str(batch_count)),
        ("Date d'export", "—"),  # À remplacer par date réelle si nécessaire
    ]
    row = 2
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10)
        row += 1

    # ── En-têtes colonnes (ajout de la colonne Facture) ─────────────────────
    headers = ["Facture", "Page PDF", "Page Doc.", "Type de contrat", "N° d'Appel",
               "Articles Mensuels", "Articles Ponctuels", "Total Contrat (MAD)"]
    header_row = row + 1
    thin = Side(style="thin", color=COLORS['gris_bord'])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=10)
        c.fill = PatternFill("solid", fgColor=COLORS['rouge'])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Données pour toutes les factures ─────────────────────────────────────
    current_row = header_row
    grand_total = 0
    
    for invoice_idx, invoice in enumerate(invoices):
        # Ajouter un séparateur de facture (optionnel)
        if invoice_idx > 0:
            current_row += 1
            separator_row = current_row
            ws.merge_cells(f"A{separator_row}:H{separator_row}")
            sep_cell = ws.cell(row=separator_row, column=1, 
                               value=f"——— Facture {invoice.invoice_number or invoice_idx+1} ———")
            sep_cell.font = Font(name="Arial", bold=True, size=10, italic=True)
            sep_cell.fill = PatternFill("solid", fgColor=COLORS['gris'])
            sep_cell.alignment = Alignment(horizontal="center")
            current_row = separator_row
        
        # Afficher les infos de la facture
        current_row += 1
        info_row = current_row
        invoice_info = [
            ("N° Facture", invoice.invoice_number or "—"),
            ("Date", invoice.invoice_date or "—"),
            ("Total", f"{invoice.total:,.2f} MAD"),
            ("Contrats", str(len(invoice.contracts))),
        ]
        
        col_offset = 0
        for label, val in invoice_info:
            ws.cell(row=info_row, column=1 + col_offset, value=label).font = Font(name="Arial", bold=True, size=9)
            ws.cell(row=info_row, column=2 + col_offset, value=val).font = Font(name="Arial", size=9)
            col_offset += 2
        
        current_row = info_row
        
        # Ajouter les contrats de cette facture
        for i, c_obj in enumerate(invoice.contracts):
            data_row = current_row + 1 + i
            values = [
                invoice.invoice_number or f"Facture {invoice_idx+1}",
                c_obj.page_number,
                c_obj.document_page or "—",
                c_obj.contract_type,
                c_obj.phone_number or "—",
                len(c_obj.frais_mensuels),
                len(c_obj.frais_ponctuels),
                c_obj.total_contrat,
            ]
            fill_color = COLORS['gris'] if (data_row - header_row - 1) % 2 == 0 else COLORS['blanc']
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = border
                if col_idx == 8:  # Colonne Total Contrat
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                    grand_total += val if isinstance(val, (int, float)) else 0
                elif col_idx == 1:  # Colonne Facture
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="center")
        
        current_row += len(invoice.contracts)

    # ── Ligne Grand Total ────────────────────────────────────────────────────
    total_row = current_row + 1
    ws.merge_cells(f"A{total_row}:G{total_row}")
    t = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    t.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=11)
    t.fill = PatternFill("solid", fgColor=COLORS['noir'])
    t.alignment = Alignment(horizontal="right")
    total_val = ws.cell(row=total_row, column=8, value=grand_total)
    total_val.font = Font(name="Arial", bold=True, color=COLORS['blanc'], size=11)
    total_val.fill = PatternFill("solid", fgColor=COLORS['noir'])
    total_val.number_format = "#,##0.00"
    total_val.alignment = Alignment(horizontal="right")

    # ── Largeurs colonnes ───────────────────────────────────────────────────
    col_widths = [15, 10, 10, 28, 18, 18, 18, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()