"""Interface web Streamlit pour l'extracteur de factures IAM YAZAKI.

Lancement :
    streamlit run app.py
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from invoice_parser import parse_invoice, InvoiceData
from expenses import calculate_expenses, calculate_expenses_from_db
import database as db


@st.cache_resource(show_spinner=False)
def load_ocr_engine():
    """Charge l'OCR engine une seule fois (singleton mis en cache par Streamlit)."""
    try:
        from ocr_engine import OCREngine
        return OCREngine(), None
    except Exception as e:
        return None, str(e)


ROOT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = ROOT_DIR / "config.json"
DATA_DIR = ROOT_DIR / "data"
MASTER_FILE = DATA_DIR / "invoices.json"

db.init_db()


# ---------------------------------------------------------------- utilitaires
def load_config() -> dict:
    if not CONFIG_PATH.exists():
        return {"plants": {}}
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return cleaned or "facture"


def append_to_master(entry: dict) -> Path:
    DATA_DIR.mkdir(exist_ok=True)
    if MASTER_FILE.exists():
        try:
            master = json.loads(MASTER_FILE.read_text(encoding="utf-8"))
            if not isinstance(master, list):
                master = []
        except json.JSONDecodeError:
            master = []
    else:
        master = []
    master.append(entry)
    MASTER_FILE.write_text(
        json.dumps(master, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return MASTER_FILE


def list_saved_invoices() -> list[Path]:
    if not DATA_DIR.exists():
        return []
    return sorted(
        [p for p in DATA_DIR.glob("*.json") if p.name != "invoices.json"],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )


def fmt_money(value: float) -> str:
    return f"{value:,.2f} MAD".replace(",", " ")


def contracts_dataframe(invoice: InvoiceData) -> pd.DataFrame:
    rows = []
    for c in invoice.contracts:
        rows.append(
            {
                "Page PDF": c.page_number,
                "Page Doc.": c.document_page or "—",
                "Type de contrat": c.contract_type,
                "N° d'Appel": c.phone_number or "—",
                "Articles (Mensuels)": len(c.frais_mensuels),
                "Articles (Ponctuels)": len(c.frais_ponctuels),
                "Total Contrat (MAD)": c.total_contrat,
            }
        )
    return pd.DataFrame(rows)


def contracts_to_excel(invoice: InvoiceData, plant: str = "", dept: str = "", project: str = "") -> bytes:
    """Génère un fichier Excel avec le détail des contrats."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Détail des contrats"

    # Couleurs YAZAKI
    rouge = "DC2626"
    noir = "0A0A0A"
    blanc = "FFFFFF"
    gris = "F5F5F5"

    # ── En-tête YAZAKI ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    cell_title = ws["A1"]
    cell_title.value = "YAZAKI · Détail des contrats IAM"
    cell_title.font = Font(name="Arial", bold=True, size=14, color=blanc)
    cell_title.fill = PatternFill("solid", fgColor=noir)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Meta-info
    meta = [
        ("Usine", plant or "—"),
        ("Département", dept or "—"),
        ("Projet", project or "—"),
        ("N° Facture", invoice.invoice_number or "—"),
        ("Date Facture", invoice.invoice_date or "—"),
        ("Période", f"{invoice.period_start or '—'} → {invoice.period_end or '—'}"),
        ("Total (MAD)", f"{invoice.total:,.2f}"),
        ("Nb contrats", str(len(invoice.contracts))),
    ]
    row = 2
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10)
        row += 1

    # ── En-têtes colonnes ───────────────────────────────────────────────────
    headers = ["Page PDF", "Page Doc.", "Type de contrat", "N° d'Appel",
               "Articles Mensuels", "Articles Ponctuels", "Total Contrat (MAD)"]
    header_row = row + 1
    thin = Side(style="thin", color="E5E5E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, color=blanc, size=10)
        c.fill = PatternFill("solid", fgColor=rouge)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Données ─────────────────────────────────────────────────────────────
    for i, c_obj in enumerate(invoice.contracts):
        data_row = header_row + 1 + i
        values = [
            c_obj.page_number,
            c_obj.document_page or "—",
            c_obj.contract_type,
            c_obj.phone_number or "—",
            len(c_obj.frais_mensuels),
            len(c_obj.frais_ponctuels),
            c_obj.total_contrat,
        ]
        fill_color = gris if i % 2 == 0 else blanc
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            if col_idx == 7:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    # ── Ligne Total ─────────────────────────────────────────────────────────
    total_row = header_row + 1 + len(invoice.contracts)
    ws.merge_cells(f"A{total_row}:F{total_row}")
    t = ws.cell(row=total_row, column=1, value="TOTAL")
    t.font = Font(name="Arial", bold=True, color=blanc, size=10)
    t.fill = PatternFill("solid", fgColor=noir)
    t.alignment = Alignment(horizontal="right")
    total_val = ws.cell(row=total_row, column=7,
                        value=f"=SUM(G{header_row+1}:G{total_row-1})")
    total_val.font = Font(name="Arial", bold=True, color=blanc, size=10)
    total_val.fill = PatternFill("solid", fgColor=noir)
    total_val.number_format = "#,##0.00"
    total_val.alignment = Alignment(horizontal="right")

    # ── Largeurs colonnes ───────────────────────────────────────────────────
    col_widths = [10, 10, 28, 18, 18, 18, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize(s: str) -> str:
    """Normalise une chaîne : minuscules, sans accents, sans tirets/espaces."""
    import unicodedata
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[\s\-_/]+", "", s)


def _load_budget_df(file_bytes: bytes) -> pd.DataFrame:
    """Charge le fichier Excel en détectant automatiquement la ligne d'en-tête.
    Supporte les fichiers où la 1ère ligne est Column1/Column2... et
    la 2ème ligne contient les vrais noms de colonnes (usine, budget...).
    """
    # Lecture brute sans en-tête
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)

    # Chercher la ligne qui contient les vrais en-têtes
    for i, row in df_raw.iterrows():
        row_norm = [_normalize(str(v)) for v in row.values]
        if any(k in row_norm for k in ("usine", "plant", "budget", "departement")):
            df_raw.columns = [str(v).strip() for v in df_raw.iloc[i].values]
            return df_raw.iloc[i + 1:].reset_index(drop=True)

    # Fallback : header sur la 1ère ligne
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=0)


def read_budget_from_excel(file_bytes: bytes, plant: str, dept: str, project: str) -> dict | None:
    """Lit le budget ET le cost center depuis un fichier Excel.
    Matching insensible aux accents, majuscules et espaces."""
    try:
        df = _load_budget_df(file_bytes)
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {}
        for col in df.columns:
            n = _normalize(col)
            if n in ("usine", "plant", "factory"): col_map["usine"] = col
            elif n in ("departement", "department", "dept"): col_map["dept"] = col
            elif n in ("projet", "project"): col_map["projet"] = col
            elif n in ("budget", "montant", "amount"): col_map["budget"] = col
            elif n in ("costcenterid", "costcenter", "centredecoût", "cc"): col_map["cc"] = col
        if not col_map.get("budget"):
            return None

        # Matching normalisé projet+usine+dept
        mask = pd.Series([True] * len(df))
        if col_map.get("usine") and plant:
            mask &= df[col_map["usine"]].astype(str).apply(_normalize) == _normalize(plant)
        if col_map.get("dept") and dept:
            mask &= df[col_map["dept"]].astype(str).apply(_normalize) == _normalize(dept)
        if col_map.get("projet") and project:
            mask &= df[col_map["projet"]].astype(str).apply(_normalize) == _normalize(project)
        matched = df[mask]

        # Fallback : matching sans le projet (noms différents usine↔excel)
        if matched.empty and (plant or dept):
            mask2 = pd.Series([True] * len(df))
            if col_map.get("usine") and plant:
                mask2 &= df[col_map["usine"]].astype(str).apply(_normalize) == _normalize(plant)
            if col_map.get("dept") and dept:
                mask2 &= df[col_map["dept"]].astype(str).apply(_normalize) == _normalize(dept)
            matched = df[mask2]

        if matched.empty:
            return None

        row_data = matched.iloc[0]
        raw_budget = str(row_data[col_map["budget"]]).replace(",", ".").replace(" ", "")
        budget_val = float(raw_budget)
        cc_val = None
        if col_map.get("cc"):
            cc_raw = row_data.get(col_map["cc"])
            cc_val = str(cc_raw).strip() if cc_raw is not None and str(cc_raw) != "nan" else None
        found_project = str(row_data.get(col_map.get("projet", ""), "")).strip() \
            if col_map.get("projet") else project
        return {"budget": budget_val, "cost_center_id": cc_val, "matched_project": found_project}
    except Exception:
        return None
    rows = []
    for c in invoice.contracts:
        rows.append(
            {
                "Page PDF": c.page_number,
                "Page Doc.": c.document_page or "—",
                "Type de contrat": c.contract_type,
                "N° d'Appel": c.phone_number or "—",
                "Articles (Mensuels)": len(c.frais_mensuels),
                "Articles (Ponctuels)": len(c.frais_ponctuels),
                "Total Contrat (MAD)": c.total_contrat,
            }
        )
    return pd.DataFrame(rows)


# --------------------------------------------------------------- configuration
st.set_page_config(
    page_title="YAZAKI · Extracteur de factures IAM",
    page_icon="🟥",
    layout="wide",
    initial_sidebar_state="expanded",
)
logo_path = ROOT_DIR / "public" / "logo.png"
if logo_path.exists():
    st.logo(str(logo_path), size="large")

# ------------------------------------------------------------------------ CSS
# Palette : noir (#0a0a0a) / blanc (#ffffff) / rouge (#dc2626)
st.markdown(
    """
    <style>
      :root {
        --rouge: #dc2626;
        --rouge-fonce: #991b1b;
        --noir: #0a0a0a;
        --blanc: #ffffff;
        --gris-clair: #f5f5f5;
        --gris-bord: #e5e5e5;
      }

      /* Conteneur principal */
      .block-container {
        padding-top: 1.6rem;
        padding-bottom: 2rem;
        background: var(--blanc);
      }
      .stApp { background: var(--blanc); color: var(--noir); }

      h1, h2, h3, h4 {
        color: var(--noir) !important;
        letter-spacing: -.01em;
        font-weight: 700;
      }
      h1 {
        border-bottom: 3px solid var(--rouge);
        padding-bottom: 0.4rem;
        display: inline-block;
      }

      /* Barre latérale : fond noir, texte blanc */
      [data-testid="stSidebar"] {
        background: var(--noir) !important;
      }
      [data-testid="stSidebar"] * {
        color: var(--blanc) !important;
      }
      [data-testid="stSidebar"] h1,
      [data-testid="stSidebar"] h2,
      [data-testid="stSidebar"] h3 {
        color: var(--blanc) !important;
        border-bottom: 2px solid var(--rouge);
        padding-bottom: .35rem;
      }
      /* Selectbox dans la sidebar : fond blanc, texte noir */
      [data-testid="stSidebar"] [data-baseweb="select"] > div {
        background: var(--blanc) !important;
        color: var(--noir) !important;
        border: 1px solid var(--rouge) !important;
        border-radius: 8px !important;
      }
      [data-testid="stSidebar"] [data-baseweb="select"] * {
        color: var(--noir) !important;
      }
      [data-testid="stSidebar"] label {
        color: var(--blanc) !important;
        font-weight: 600 !important;
      }

      /* Cartes / metric */
      div[data-testid="stMetric"] {
        background: var(--blanc);
        border: 1px solid var(--gris-bord);
        border-left: 4px solid var(--rouge);
        padding: 14px 16px;
        border-radius: 10px;
      }
      div[data-testid="stMetricLabel"] { color: var(--noir) !important; }
      div[data-testid="stMetricValue"] {
        color: var(--noir) !important;
        font-weight: 700;
      }

      /* Onglets */
      .stTabs [data-baseweb="tab-list"] {
        gap: 0.4rem;
        border-bottom: 2px solid var(--noir);
      }
      .stTabs [data-baseweb="tab"] {
        padding: 0.6rem 1.1rem;
        border-radius: 10px 10px 0 0;
        background: var(--gris-clair);
        color: var(--noir) !important;
        border: 1px solid var(--gris-bord);
        font-weight: 600;
      }
      .stTabs [aria-selected="true"] {
        background: var(--rouge) !important;
        color: var(--blanc) !important;
        border-color: var(--rouge) !important;
      }
      .stTabs [aria-selected="true"] * { color: var(--blanc) !important; }

      /* Boutons : tous lisibles (corrige le blanc sur blanc) */
      .stButton > button,
      .stDownloadButton > button,
      .stForm button {
        background: var(--noir) !important;
        color: var(--blanc) !important;
        border: 1px solid var(--noir) !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.55rem 1.1rem !important;
        transition: all 0.15s ease !important;
      }
      .stButton > button:hover,
      .stDownloadButton > button:hover {
        background: var(--rouge) !important;
        color: var(--blanc) !important;
        border-color: var(--rouge) !important;
        transform: translateY(-1px);
      }
      .stButton > button[kind="primary"],
      .stForm button[kind="primary"] {
        background: var(--rouge) !important;
        color: var(--blanc) !important;
        border: 1px solid var(--rouge) !important;
      }
      .stButton > button[kind="primary"]:hover {
        background: var(--rouge-fonce) !important;
        border-color: var(--rouge-fonce) !important;
      }
      .stButton > button:disabled,
      .stDownloadButton > button:disabled {
        background: var(--gris-clair) !important;
        color: #a3a3a3 !important;
        border-color: var(--gris-bord) !important;
      }

      /* Champs de saisie */
      .stTextInput input, .stNumberInput input,
      [data-baseweb="select"] > div {
        border: 1px solid var(--gris-bord) !important;
        border-radius: 8px !important;
      }
      .stTextInput input:focus, .stNumberInput input:focus {
        border-color: var(--rouge) !important;
        box-shadow: 0 0 0 1px var(--rouge) !important;
      }

      /* File uploader */
      [data-testid="stFileUploader"] section {
        background: var(--blanc) !important;
        border: 2px dashed var(--rouge) !important;
        border-radius: 10px !important;
      }
      [data-testid="stFileUploader"] button {
        background: var(--rouge) !important;
        color: var(--blanc) !important;
        border: 1px solid var(--rouge) !important;
      }

      /* Radio horizontal : lisible */
      div[role="radiogroup"] label {
        background: var(--blanc);
        border: 1px solid var(--gris-bord);
        padding: 6px 12px;
        border-radius: 8px;
        margin-right: 6px;
        color: var(--noir) !important;
        font-weight: 500;
      }
      div[role="radiogroup"] label:has(input:checked) {
        background: var(--rouge) !important;
        border-color: var(--rouge) !important;
      }
      div[role="radiogroup"] label:has(input:checked) * {
        color: var(--blanc) !important;
      }

      /* Cartes infos custom */
      .carte-info {
        padding: 1rem 1.2rem;
        border-radius: 12px;
        background: var(--blanc);
        border: 1px solid var(--gris-bord);
        border-top: 3px solid var(--rouge);
      }
      .carte-info b { color: var(--noir); }
      .carte-info span { color: #525252; }

      /* Alertes */
      div[data-testid="stAlertContainer"] {
        border-radius: 10px;
      }

      /* Tableau / dataframe */
      [data-testid="stDataFrame"] {
        border: 1px solid var(--gris-bord);
        border-radius: 10px;
        overflow: hidden;
      }

      /* Liens */
      a { color: var(--rouge) !important; }
      a:hover { color: var(--rouge-fonce) !important; }
    </style>
    """,
    unsafe_allow_html=True,
)


# --------------------------------------------------------------------- en-tête
st.title("YAZAKI · Extracteur de factures IAM")
st.caption(
    "Téléchargez une facture Maroc Telecom multipage → catégorisez-la par "
    "Usine / Département / Projet → enregistrez en JSON → calculez "
    "Budget vs. Dépenses."
)


if "invoice" not in st.session_state:
    st.session_state.invoice = None
if "uploaded_name" not in st.session_state:
    st.session_state.uploaded_name = None
if "auto_saved" not in st.session_state:
    st.session_state.auto_saved = False
if "cost_center_id" not in st.session_state:
    st.session_state.cost_center_id = None
if "pdf_bytes" not in st.session_state:
    st.session_state.pdf_bytes = None
if "pdf_file_key" not in st.session_state:
    st.session_state.pdf_file_key = None


config = load_config()
plants_data: dict = config.get("plants", {})


# ------------------------------------------------------------- barre latérale
with st.sidebar:
    st.header("Catégorisation")
    st.caption("Choisissez l'Usine, le Département et le Projet pour cette facture.")

    plant = st.selectbox(
        "Usine",
        options=[""] + list(plants_data.keys()),
        format_func=lambda x: x if x else "— Choisir une usine —",
        key="plant",
    )

    departments = list(plants_data.get(plant, {}).keys()) if plant else []
    dept = st.selectbox(
        "Département",
        options=[""] + departments,
        format_func=lambda x: x if x else "— Choisir un département —",
        disabled=not plant,
        key="department",
    )

    projects = plants_data.get(plant, {}).get(dept, []) if (plant and dept) else []
    project = st.selectbox(
        "Projet",
        options=[""] + list(projects),
        format_func=lambda x: x if x else "— Choisir un projet —",
        disabled=not (plant and dept),
        key="project",
    )

    st.divider()
    st.caption("Modifiez les options dans `config.json`, puis actualisez la page.")


# --------------------------------------------------------------------- onglets
tab_extract, tab_expenses, tab_history = st.tabs(
    ["Extraire la facture", "Calculer les dépenses", "Historique"]
)


# ============================================================== ONGLET EXTRAIRE
with tab_extract:
    upload_col, info_col = st.columns([2, 1])

    with upload_col:
        uploaded = st.file_uploader(
            "Téléchargez une facture IAM multipage (PDF)",
            type=["pdf"],
            accept_multiple_files=False,
        )

    with info_col:
        st.markdown(
            """
            <div class='carte-info'>
              <b>Calcul du Total</b><br/>
              <span>
              Somme du <code>TOTAL CONTRAT</code> de chaque page de contrat —
              c.-à-d. toutes les lignes <i>sauf la première</i>
              (récapitulatif global), projetées sur la colonne
              <i>Total Contrat</i>.
              </span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Stocker les bytes du PDF dès l'upload (une seule lecture) ────────
    if uploaded is not None:
        file_key = f"pdf_bytes_{uploaded.name}_{uploaded.size}"
        if st.session_state.get("pdf_file_key") != file_key:
            st.session_state.pdf_bytes = uploaded.read()
            st.session_state.pdf_file_key = file_key
        pdf_bytes = st.session_state.pdf_bytes
    else:
        pdf_bytes = None
        st.session_state.pdf_bytes = None
        st.session_state.pdf_file_key = None

    # ── Détection automatique PDF scanné + toggle OCR ───────────────────
    ocr_available = True
    ocr_import_error = None
    try:
        import easyocr  # noqa: F401
        import fitz     # noqa: F401
    except ImportError as e:
        ocr_available = False
        ocr_import_error = str(e)

    use_ocr = False
    if pdf_bytes is not None:
        try:
            import fitz
            pdf_preview = fitz.open(stream=pdf_bytes, filetype="pdf")
            is_scanned = all(
                len(pdf_preview[i].get_text().strip()) < 80
                for i in range(min(3, len(pdf_preview)))
            )
            pdf_preview.close()
        except Exception:
            is_scanned = False

        if is_scanned and ocr_available:
            st.warning(
                "⚠️ Ce PDF semble **scanné** (pas de texte extractible). "
                "Le mode OCR a été activé automatiquement."
            )
            use_ocr = True
        elif is_scanned and not ocr_available:
            st.error(
                f"⚠️ Ce PDF semble scanné mais OCR non disponible : `{ocr_import_error}`. "
                "Lancez `pip install easyocr PyMuPDF` puis redémarrez."
            )
        elif not ocr_available:
            st.caption(f"ℹ️ OCR non disponible (`{ocr_import_error}`). Installez easyocr et PyMuPDF.")
        elif ocr_available:
            use_ocr = st.toggle(
                "🔍 Forcer le mode OCR (PDF scanné)",
                value=False,
                help="Activez si des pages sont manquantes ou vides dans le résultat.",
            )

    extract_clicked = st.button(
        "Extraire les données",
        type="primary",
        disabled=pdf_bytes is None,
        use_container_width=False,
    )

    if extract_clicked and pdf_bytes is not None:
        ocr_engine = None
        if use_ocr:
            with st.spinner("⏳ Chargement du moteur OCR (première fois : téléchargement du modèle ~500 Mo)…"):
                ocr_engine, ocr_error = load_ocr_engine()
            if ocr_engine is None:
                st.error(f"❌ Impossible de charger EasyOCR : `{ocr_error}`")
                st.info("💡 Essayez dans le terminal :\n```\npip uninstall easyocr -y\npip install easyocr\n```")
                st.stop()

        spinner_msg = (
            "🔍 OCR en cours… (peut prendre 1-2 min pour un PDF de 80+ pages)"
            if use_ocr
            else "Analyse du PDF en cours…"
        )
        with st.spinner(spinner_msg):
            progress_bar = st.progress(0, text="Démarrage…")
            status_text  = st.empty()

            def on_progress(current, total, status):
                pct = int((current / total) * 100) if total else 0
                progress_bar.progress(pct, text=f"{status} ({pct}%)")
                status_text.caption(status)

            try:
                buf = io.BytesIO(pdf_bytes)
                invoice = parse_invoice(buf, ocr_engine=ocr_engine,
                                        progress_callback=on_progress)
                invoice.source_file = uploaded.name if uploaded else "inconnu"
                st.session_state.invoice = invoice
                st.session_state.uploaded_name = invoice.source_file

                progress_bar.progress(100, text="✅ Extraction terminée !")
                status_text.empty()
                ocr_label = " (via OCR)" if use_ocr else ""

                if len(invoice.contracts) == 0:
                    st.warning(
                        f"⚠️ 0 contrats extraits{ocr_label}. "
                        + ("Le texte OCR n'a pas pu être analysé — vérifiez la qualité du scan."
                           if use_ocr else
                           "Vérifiez que le PDF est bien une facture IAM.")
                    )
                else:
                    st.success(
                        f"✅ {len(invoice.contracts)} pages de contrat extraites{ocr_label} "
                        f"depuis « {invoice.source_file} »."
                    )
            except Exception as exc:
                progress_bar.empty()
                status_text.empty()
                st.session_state.invoice = None
                st.error(f"Échec de l'analyse du PDF : {exc}")

    invoice: InvoiceData | None = st.session_state.invoice

    if invoice is not None:
        st.subheader("Récapitulatif")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("N° Facture", invoice.invoice_number or "—")
        m2.metric("Date Facture", invoice.invoice_date or "—")
        m3.metric("Nombre de contrats", len(invoice.contracts))
        m4.metric("Total (MAD)", fmt_money(invoice.total))

        m5, m6, m7, m8 = st.columns(4)
        m5.metric("Début période", invoice.period_start or "—")
        m6.metric("Fin période", invoice.period_end or "—")
        m7.metric(
            "Montant HT global",
            fmt_money(invoice.global_summary.montant_ht),
            help="Lu sur la page de récapitulatif global (vérification croisée).",
        )
        m8.metric(
            "Montant TTC global",
            fmt_money(invoice.global_summary.montant_ttc),
        )

        delta = round(invoice.total - invoice.global_summary.montant_ht, 2)
        if invoice.global_summary.montant_ht:
            if abs(delta) < 0.05:
                st.success(
                    f"Le total calculé correspond exactement au Montant HT "
                    f"({fmt_money(invoice.total)})."
                )
            else:
                st.warning(
                    f"Écart entre le total calculé et le Montant HT global : "
                    f"{fmt_money(delta)}"
                )

        st.subheader("Détail des contrats")
        df = contracts_dataframe(invoice)
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Total Contrat (MAD)": st.column_config.NumberColumn(format="%.2f"),
            },
        )

        # ── Téléchargement Excel des contrats ────────────────────────────
        excel_bytes = contracts_to_excel(invoice, plant, dept, project)
        st.download_button(
            "📥 Télécharger le détail des contrats (Excel)",
            data=excel_bytes,
            file_name=safe_filename(
                f"{plant or 'usine'}_{dept or 'dept'}_{project or 'projet'}"
                f"_{invoice.invoice_number or 'contrats'}"
            ) + "_contrats.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

        with st.expander("Afficher le JSON complet extrait"):
            st.json(invoice.to_dict(), expanded=False)

        st.divider()
        st.subheader("Enregistrer cette facture")

        record = {
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "plant": plant,
            "department": dept,
            "project": project,
            "cost_center_id": st.session_state.get("cost_center_id"),
            **invoice.to_dict(),
        }
        json_bytes = json.dumps(record, indent=2, ensure_ascii=False).encode("utf-8")

        # ── Auto-sauvegarde si catégorie complète ────────────────────────
        if plant and dept and project:
            invoice_key = f"{plant}_{dept}_{project}_{invoice.invoice_number}"
            if not st.session_state.get(f"saved_{invoice_key}"):
                DATA_DIR.mkdir(exist_ok=True)
                invoice_num = invoice.invoice_number or "facture"
                filename = safe_filename(f"{plant}_{dept}_{project}_{invoice_num}") + ".json"
                per_file = DATA_DIR / filename
                per_file.write_text(
                    json.dumps(record, indent=2, ensure_ascii=False), encoding="utf-8"
                )
                append_to_master(record)
                try:
                    invoice_id = db.save_invoice(record)
                    st.session_state[f"saved_{invoice_key}"] = True
                    st.success(
                        f"✅ Sauvegarde automatique effectuée — SQLite (id={invoice_id}), "
                        f"fichier `{filename}` créé."
                    )
                except Exception as exc:  # noqa: BLE001
                    st.warning(f"Fichiers JSON enregistrés, mais SQLite a échoué : {exc}")
            else:
                st.success(f"✅ Déjà sauvegardé pour {plant} / {dept} / {project}.")

            # ── Téléchargement JSON automatique (toujours disponible) ────
            st.download_button(
                "📥 Télécharger le JSON",
                data=json_bytes,
                file_name=safe_filename(
                    f"{plant}_{dept}_{project}_{invoice.invoice_number or 'facture'}"
                ) + ".json",
                mime="application/json",
                use_container_width=False,
            )
        else:
            st.info(
                "Choisissez **Usine → Département → Projet** dans la barre "
                "latérale pour activer la sauvegarde automatique."
            )
            st.download_button(
                "📥 Télécharger le JSON (sans catégorie)",
                data=json_bytes,
                file_name="facture.json",
                mime="application/json",
                use_container_width=False,
            )
    else:
        st.info(
            "Téléchargez un PDF puis cliquez sur **Extraire les données** "
            "pour commencer."
        )


# ============================================================ ONGLET DÉPENSES
with tab_expenses:
    st.subheader("Calculer le budget restant")
    st.caption(
        "Résultat = Budget − Total. Le Total peut provenir d'une seule "
        "facture ou être agrégé depuis la base SQLite."
    )

    src = st.radio(
        "Source",
        (
            "Base SQLite (agrégat par Usine/Dépt./Projet)",
            "Facture enregistrée (SQLite)",
            "Fichier JSON enregistré (./data/)",
            "Téléverser un fichier JSON",
            "Facture en cours d'extraction",
        ),
        horizontal=False,
    )

    json_data: dict | None = None
    src_label = ""
    db_aggregate: dict | None = None

    if src == "Base SQLite (agrégat par Usine/Dépt./Projet)":
        plants_db = db.distinct_values("plant")
        if not plants_db:
            st.info(
                "Aucune facture en base. Enregistrez-en une depuis l'onglet "
                "**Extraire la facture**."
            )
        else:
            f1, f2, f3 = st.columns(3)
            with f1:
                fplant = st.selectbox(
                    "Usine (filtre)",
                    options=["(toutes)"] + plants_db,
                    key="f_plant",
                )
            with f2:
                deps = (
                    [d for d in db.distinct_values("department")]
                    if fplant != "(toutes)"
                    else db.distinct_values("department")
                )
                fdept = st.selectbox(
                    "Département (filtre)",
                    options=["(tous)"] + deps,
                    key="f_dept",
                )
            with f3:
                projs = db.distinct_values("project")
                fproj = st.selectbox(
                    "Projet (filtre)",
                    options=["(tous)"] + projs,
                    key="f_proj",
                )
            db_aggregate = db.aggregate_total(
                plant=None if fplant == "(toutes)" else fplant,
                department=None if fdept == "(tous)" else fdept,
                project=None if fproj == "(tous)" else fproj,
            )
            src_label = (
                f"SQLite — Usine: {fplant} | Dépt: {fdept} | Projet: {fproj} "
                f"({db_aggregate['invoices_count']} facture(s), "
                f"{db_aggregate['contracts_count']} contrats)"
            )

    elif src == "Facture enregistrée (SQLite)":
        rows = db.list_invoices()
        if not rows:
            st.info("Aucune facture en base.")
        else:
            options = {
                f"#{r['id']} — {r['plant']} / {r['department']} / "
                f"{r['project']} — Facture {r['invoice_number']} "
                f"({fmt_money(r['total'])})": r["id"]
                for r in rows
            }
            choice = st.selectbox("Facture en base", list(options.keys()))
            if choice:
                json_data = db.get_invoice_raw(options[choice])
                src_label = choice

    elif src == "Fichier JSON enregistré (./data/)":
        saved = list_saved_invoices()
        if not saved:
            st.info("Aucun fichier JSON dans ./data/.")
        else:
            choice = st.selectbox(
                "Fichier JSON", saved, format_func=lambda p: p.name
            )
            if choice:
                json_data = json.loads(Path(choice).read_text(encoding="utf-8"))
                src_label = choice.name

    elif src == "Téléverser un fichier JSON":
        up = st.file_uploader("Téléverser un JSON", type=["json"], key="exp_upload")
        if up is not None:
            json_data = json.loads(up.read().decode("utf-8"))
            src_label = up.name

    else:  # facture en cours
        if st.session_state.invoice is None:
            st.info(
                "Extrayez d'abord une facture dans l'onglet "
                "**Extraire la facture**."
            )
        else:
            json_data = st.session_state.invoice.to_dict()
            src_label = st.session_state.uploaded_name or "facture en cours"

    # ── Source du budget ─────────────────────────────────────────────────────
    st.markdown("**Source du budget**")
    budget_src = st.radio(
        "Source du budget",
        ("Saisir manuellement", "Importer depuis Excel (BUDGET.xlsx)"),
        horizontal=True,
        label_visibility="collapsed",
    )

    budget_excel_value: float | None = None
    if budget_src == "Importer depuis Excel (BUDGET.xlsx)":
        up_budget = st.file_uploader(
            "Téléverser le fichier BUDGET.xlsx",
            type=["xlsx", "xls"],
            key="budget_excel",
        )
        if up_budget is not None:
            budget_bytes = up_budget.read()

            # ── Essai de matching automatique (sidebar) ──────────────────
            excel_result = read_budget_from_excel(budget_bytes, plant, dept, project)

            if excel_result is not None:
                budget_excel_value = excel_result["budget"]
                st.session_state.cost_center_id = excel_result["cost_center_id"]
                cc_display = excel_result["cost_center_id"] or "—"
                matched_proj = excel_result.get("matched_project", project)
                st.success(
                    f"✅ Budget lu automatiquement : **{fmt_money(budget_excel_value)}** "
                    f"| Cost Center : **{cc_display}** "
                    f"— {plant or '?'} / {dept or '?'} / {matched_proj or '?'}"
                )
            else:
                # ── Fallback : afficher toutes les lignes du fichier ─────
                st.info(
                    "Aucune correspondance automatique trouvée. "
                    "Sélectionnez manuellement la ligne budget ci-dessous :"
                )
                try:
                    df_budget = pd.read_excel(io.BytesIO(budget_bytes), sheet_name=0)
                    # Normalise les colonnes
                    df_budget.columns = [str(c).strip() for c in df_budget.columns]
                    # Ignore la première ligne si c'est un doublon d'en-têtes
                    if df_budget.iloc[0].astype(str).str.lower().tolist() == \
                       [c.lower() for c in df_budget.columns]:
                        df_budget = df_budget.iloc[1:].reset_index(drop=True)

                    # Construire les options d'affichage
                    cols = df_budget.columns.tolist()
                    options_map = {}
                    for _, row in df_budget.iterrows():
                        label = " | ".join(
                            f"{c}: {row[c]}" for c in cols
                        )
                        options_map[label] = row

                    if options_map:
                        chosen_label = st.selectbox(
                            "Choisir la ligne budget :",
                            list(options_map.keys()),
                            key="budget_row_select",
                        )
                        chosen_row = options_map[chosen_label]

                        # Trouver la colonne budget
                        budget_col = next(
                            (c for c in cols if _normalize(c) in
                             ("budget", "montant", "amount")), None
                        )
                        cc_col = next(
                            (c for c in cols if _normalize(c) in
                             ("costcenterid", "costcenter", "centredecoût", "cc")), None
                        )

                        if budget_col:
                            raw_b = str(chosen_row[budget_col]).replace(",", ".").replace(" ", "")
                            budget_excel_value = float(raw_b)
                            cc_val = str(chosen_row[cc_col]).strip() if cc_col else None
                            if cc_val == "nan": cc_val = None
                            st.session_state.cost_center_id = cc_val
                            st.success(
                                f"✅ Budget sélectionné : **{fmt_money(budget_excel_value)}** "
                                f"| Cost Center : **{cc_val or '—'}**"
                            )
                except Exception as exc:
                    st.error(f"Erreur de lecture du fichier Excel : {exc}")

    default_budget = budget_excel_value if budget_excel_value is not None else 60000.0
    budget = st.number_input(
        "Budget (MAD)",
        min_value=0.0,
        value=default_budget,
        step=500.0,
        format="%.2f",
        disabled=(budget_src == "Importer depuis Excel (BUDGET.xlsx)" and budget_excel_value is not None),
    )

    if db_aggregate is not None:
        total = db_aggregate["total"]
        result = {
            "budget": float(budget),
            "total": total,
            "expenses": round(float(budget) - total, 2),
        }
    elif json_data is not None:
        result = calculate_expenses(json_data, budget)
    else:
        result = None

    if result is not None:
        c1, c2, c3 = st.columns(3)
        c1.metric("Budget", fmt_money(result["budget"]))
        c2.metric("Total", fmt_money(result["total"]))
        delta = result["expenses"]
        c3.metric(
            "Dépenses (Budget − Total)",
            fmt_money(delta),
            delta=fmt_money(delta),
            delta_color="normal" if delta >= 0 else "inverse",
        )
        st.caption(f"Source : `{src_label}`")

        if json_data is not None:
            # ── Export Excel contrats depuis la source JSON ───────────────
            inv_tmp = None
            try:
                from invoice_parser import InvoiceData
                inv_tmp = InvoiceData.from_dict(json_data) if hasattr(InvoiceData, "from_dict") else None
            except Exception:
                pass

            if inv_tmp is None and json_data.get("contracts"):
                # Fallback : construire un DataFrame simple depuis json_data
                rows_exp = []
                for c_obj in json_data.get("contracts", []):
                    rows_exp.append({
                        "Page PDF": c_obj.get("page_number", ""),
                        "Page Doc.": c_obj.get("document_page", "—"),
                        "Type de contrat": c_obj.get("contract_type", ""),
                        "N° d'Appel": c_obj.get("phone_number", "—"),
                        "Articles Mensuels": len(c_obj.get("frais_mensuels", [])),
                        "Articles Ponctuels": len(c_obj.get("frais_ponctuels", [])),
                        "Total Contrat (MAD)": c_obj.get("total_contrat", 0),
                    })
                if rows_exp:
                    df_exp = pd.DataFrame(rows_exp)
                    buf_exp = io.BytesIO()
                    df_exp.to_excel(buf_exp, index=False, sheet_name="Contrats")
                    st.download_button(
                        "📥 Télécharger les contrats (Excel)",
                        data=buf_exp.getvalue(),
                        file_name=f"contrats_{src_label.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            with st.expander("Aperçu du JSON"):
                st.json(json_data, expanded=False)


# ========================================================== ONGLET HISTORIQUE
with tab_history:
    st.subheader("Factures enregistrées (base SQLite)")

    stats = db.db_stats()
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Factures", stats["invoices"])
    s2.metric("Contrats", stats["contracts"])
    s3.metric("Lignes (items)", stats["line_items"])
    s4.metric("Taille DB", f"{db.db_size_bytes()/1024:,.1f} Ko")

    rows = db.list_invoices()
    if not rows:
        st.info("Aucune facture en base pour l'instant.")
    else:
        df = pd.DataFrame(
            [
                {
                    "id": r["id"],
                    "Enregistrée le": r["saved_at"],
                    "Usine": r["plant"],
                    "Département": r["department"],
                    "Projet": r["project"],
                    "Cost Center": r.get("cost_center_id") or "—",
                    "N° Facture": r["invoice_number"] or "—",
                    "Date": r["invoice_date"] or "—",
                    "Période": (
                        f"{r['period_start'] or '—'} → {r['period_end'] or '—'}"
                    ),
                    "Contrats": r["contracts_count"],
                    "Total (MAD)": r["total"],
                    "Mt HT": r["montant_ht"],
                    "Mt TTC": r["montant_ttc"],
                }
                for r in rows
            ]
        )
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Total (MAD)": st.column_config.NumberColumn(format="%.2f"),
                "Mt HT": st.column_config.NumberColumn(format="%.2f"),
                "Mt TTC": st.column_config.NumberColumn(format="%.2f"),
            },
        )

        st.divider()
        st.markdown("**Détail / téléchargement / suppression**")
        for r in rows:
            label = (
                f"#{r['id']} — {r['plant']} / {r['department']} / "
                f"{r['project']} — Facture {r['invoice_number']} "
                f"({fmt_money(r['total'])})"
            )
            with st.expander(label):
                raw = db.get_invoice_raw(r["id"]) or {}
                colA, colB = st.columns([1, 1])
                with colA:
                    st.download_button(
                        "Télécharger le JSON",
                        data=json.dumps(
                            raw, indent=2, ensure_ascii=False
                        ).encode("utf-8"),
                        file_name=safe_filename(
                            f"{r['plant']}_{r['department']}_"
                            f"{r['project']}_{r['invoice_number'] or 'facture'}"
                        ) + ".json",
                        mime="application/json",
                        key=f"db_dl_{r['id']}",
                        use_container_width=True,
                    )
                with colB:
                    if st.button(
                        "Supprimer de la base",
                        key=f"db_del_{r['id']}",
                        use_container_width=True,
                    ):
                        db.delete_invoice(r["id"])
                        st.success(f"Facture #{r['id']} supprimée.")
                        st.rerun()
                st.json(raw, expanded=False)

    st.divider()
    st.markdown("**Fichiers JSON sur disque (sauvegarde) :**")
    saved = list_saved_invoices()
    if not saved:
        st.caption("Aucun fichier dans ./data/ pour l'instant.")
    for p in saved:
        with st.expander(p.name):
                content = p.read_text(encoding="utf-8")
                st.download_button(
                    "Télécharger",
                    data=content.encode("utf-8"),
                    file_name=p.name,
                    mime="application/json",
                    key=f"dl_{p.name}",
                )
                st.code(content[:2000] + ("\n..." if len(content) > 2000 else ""))